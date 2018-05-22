from statistics import mean, median, stdev
from datetime import datetime
import json

import django
django.setup()
from openpyxl import Workbook
from django.utils import timezone

from experiment.models import User, Result

class Namespace:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)

def get_all_firstresult(uid_min=0):
    ls = list(u for u in User.objects.all() if u.firstresult and u.id > uid_min)
    data_hastime = []
    data_notime = []
    for u in ls:
        rawdata = json.loads(u.firstresult.data)
        d = {k:v for k,v in rawdata.items() if k != 'mode'}
        d.update({'uid': u.id})
        if rawdata['mode'] == 'hastime':
            data_hastime.append(d)
        elif rawdata['mode'] == 'notime':
            data_notime.append(d)
    return data_hastime, data_notime

FMT = '%Y-%m-%d %H:%M %z'
BEFORE_MICHUT = "2018-05-20 17:00 +0900"
BEFORE_MICHUT_2 = "2018-05-20 20:40 +0900"
TIME = BEFORE_MICHUT_2

def getnow():
    return timezone.now().strftime(FMT)

def get_all_result_after(time=TIME):
    time = datetime.strptime(time, FMT)
    ls = list(r for r in Result.objects.all() if r.date > time)
    data_hastime = []
    data_notime = []
    for r in ls:
        rawdata = json.loads(r.data)
        if rawdata['mode'] == 'hastime':
            data_hastime.append({k:v for k,v in rawdata.items() if k != 'mode'})
        elif rawdata['mode'] == 'notime':
            data_notime.append({k:v for k,v in rawdata.items() if k != 'mode'})
    return data_hastime, data_notime

def analyze_average_final_time(data):
    avg = mean(int(d['answeredTime'][-1]) for d in data)
    med = median(int(d['answeredTime'][-1]) for d in data)
    std = stdev(int(d['answeredTime'][-1]) for d in data)
    return avg, med, std

def analyze_average_intermediate_time(data):
    avg = [mean(data[i]['answeredTime'][j] for i in range(len(data[0]))) for j in range(len(data[0]['answeredTime']))]
    med = [median(data[i]['answeredTime'][j] for i in range(len(data[0]))) for j in range(len(data[0]['answeredTime']))]
    std = [stdev(data[i]['answeredTime'][j] for i in range(len(data[0]))) for j in range(len(data[0]['answeredTime']))]
    return avg, med, std

def analyze_average_correct(data):
    avg = mean(mean(d['iscorrect']) for d in data)
    med = median(mean(d['iscorrect']) for d in data)
    std = stdev(mean(d['iscorrect']) for d in data)
    return avg, med, std

def analyze_question_time(data):
    qdict = dict()
    for d in data:
        for i, q in enumerate(d['questions']):
            has = q in qdict
            if has:
                qdict[q].append(d['answeredTime'][i] if i == 0 else d['answeredTime'][i]-d['answeredTime'][i-1])
            else:
                qdict[q] = [d['answeredTime'][i] if i == 0 else d['answeredTime'][i]-d['answeredTime'][i-1]]
    return {k:(mean(v), median(v), stdev(v) if len(v)>=2 else 0) for k,v in qdict.items()}

def export(datas, name='data.xlsx'):
    '''
    datas: Return value of get_all_firstresult(). So we can call like export(get_all_firstresult())
    name: name.
    '''
    wb = Workbook()
    wss = [wb.create_sheet("hastime"), wb.create_sheet("notime")]
    for data, ws in zip(datas, wss):
        ws['A1'] = 'uid'
        for i in range(2, 27):
            ws.cell(row=1, column=i, value='Q%d'%(i-1))
        for i, d in enumerate(data):
            ws.cell(row=4*i+2, column=1, value=d['uid'])
            for j in range(2, 27):
                ws.cell(row=4*i+2, column=j, value=d['questions'][j-2])
                ws.cell(row=4*i+3, column=j, value=d['answeredTime'][j-2])
                ws.cell(row=4*i+4, column=j, value=d['iscorrect'][j-2])
    wb.save(name)
    print('Done')

if __name__ == '__main__':
    __import__('code').interact('Sciwriting data analyze shell', local=locals())